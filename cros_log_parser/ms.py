#!/usr/bin/python
import email
import logging
import time
from datetime import datetime, timedelta
from email.parser import Parser

from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header, decode_header
from email.utils import formataddr
import zipfile
import re
import smtplib
import poplib
import os
import pandas as pd
import CONF

import Utils

logging.basicConfig(format='%(asctime)s|%(levelname)s|%(message)s', level=logging.INFO, filename='log.txt')


def move_cros_log(source_dir: str, tar_dir: str, file_name: str):
    """
    将cros_log从系统默认的下载路径转移到希望保存的路径
    :param source_dir: 原保存位置
    :param tar_dir: 期望的保存位置
    :param file_name: 文件名
    :return: None or 1
    """
    if not os.path.isdir(tar_dir):
        os.makedirs(tar_dir)
    if os.path.isfile(os.path.join(tar_dir, file_name)):
        logging.info(f" {file_name} 已存在")
        return 1
    if not os.path.isfile(os.path.join(source_dir, file_name)):
        logging.critical(f" {file_name} 不存在.")
        return 1
    with open(os.path.join(source_dir, file_name), "rb") as f:
        with open(os.path.join(tar_dir, file_name), "wb+") as nf:
            nf.write(f.read())
            logging.info(f" {file_name} 移动成功")
    # ctrx + x
    if CONF.CLEAR_SOURCE_EXCEL:
        if os.path.isfile(os.path.join(source_dir, file_name)):
            os.remove(os.path.join(source_dir, file_name))
    return 1


class Tool:
    def __init__(self):
        self.inits()
        self.res = {}

    def inits(self):

        self.zx_sum = 0  # 请求总量
        self.zx_request_failed = 0  # 请求失败
        self.zx_already_regist = 0  # 已注册用户
        self.zx_unregist = 0  # 未注册
        self.zx_ocr_failed = 0  # ocr失败
        self.zx_code_06 = 0
        self.zx_code_04 = 0
        self.zx_succ = 0
        self.zx_succ_rate = 0

        self.sx_sum = 0  # 授信总量
        self.sx_request_failed = 0  # 请求失败
        self.sx_succ = 0  # 授信成功
        self.sx_ilog_reject = 0  # 信用拒绝
        self.sx_antifraud_reject = 0  # 反欺诈拒绝
        self.sx_net_failed = 0  # 网络错误
        self.sx_succ_rate = 0  # 授信成功率

        self.tx_sum = 0  # 提现总笔数
        self.tx_request_failed = 0  # 请求失败
        self.tx_succ = 0  # 提现成功笔数
        self.tx_antifraud_reject = 0  # 反欺诈拒绝
        self.tx_ilog_reject = 0  # 风控拒绝
        self.tx_pay_failed = 0  # 支付拒绝
        self.tx_amtques = 0  # 额度问题
        self.tx_succ_rate = 0

        self.ch_flag = False

    def gen_result(self, dt: pd.DataFrame):
        """
        {"HB":[1,2,3,...],}
        :param dt: 表格数据
        :return:
        """
        for col, row in dt.iterrows():
            try:
                if row['production_code'] != dt.at[col + 1, 'production_code']:
                    self.ch_flag = True
            except KeyError as e:
                try:
                    print("换行:", f"{row['production_code']} -> f{dt.at[col + 1, 'production_code']}")
                except KeyError as e:
                    print()
            if row['log_type'] == "realnameauth":
                self.zx_sum += int(row['count'])
                if str(row['rsp_code']) == "1" or len(str(row['rsp_code'])) > 3:
                    self.zx_request_failed += int(row['count'])
                elif str(row['rsp_code']) == '3':
                    self.zx_already_regist += int(row['count'])
                self.zx_unregist = self.zx_sum - self.zx_request_failed - self.zx_already_regist
            if row['log_type'] == 'realnameauth_query':
                if str(row['rsp_code']) == "8":
                    self.zx_ocr_failed += int(row['count'])
                if str(row['rsp_code']) == "4":
                    self.zx_code_04 = int(row['count'])
                if str(row['rsp_code']) == "6":
                    self.zx_code_06 = int(row['count'])
                if str(row['rsp_code']) == "0":
                    self.zx_succ = int(row['count'])
                if self.zx_sum:
                    self.zx_succ_rate = self.zx_succ / self.zx_sum
            if row['log_type'] == "credit":
                self.sx_sum += int(row['count'])
                if row['result'] == "success" and str(row['rsp_code']) == "0":
                    self.sx_succ += int(row['count'])
                if row['err'] == "antifraud reject" and str(row['rsp_code']) == "1":
                    self.sx_antifraud_reject += int(row['count'])
                if row['result'] == "reject" and str(row['rsp_code']) == "0":
                    self.sx_ilog_reject += int(row['count'])
                if row['err'] != "antifraud reject" and str(row['rsp_code']) == "1":
                    self.sx_request_failed += int(row['count'])
                if len(str(row['rsp_code'])) > 2:
                    self.sx_request_failed += int(row['count'])
                if self.sx_sum:
                    self.sx_succ_rate = self.sx_succ / self.sx_sum
            if row['log_type'] == 'withdraw':
                if row['result'] != "review":
                    self.tx_sum += int(row['count'])
                if row['result'] == "reject" and str(row['rsp_code']) == '0':
                    self.tx_ilog_reject += int(row['count'])
                if row['err'] == "antifraud reject" and str(row['rsp_code']) == '1':
                    self.tx_antifraud_reject += int(row['count'])
                if row['err'] in CONF.REQUEST_FAILED:
                    self.tx_request_failed += int(row['count'])
                if str(row['production_code']) == "SQC" and row['result'] == "success":
                    self.tx_succ += int(row['count'])
            if row['log_type'] == 'withdraw_query':
                if row['result'] in ("BCMS4027", "BCMS3021", "BCMS3013"):
                    self.tx_amtques += int(row['count'])
                if row['production_code'] != "HB":  # 还呗的支付通道拒绝不从这个事件记录里面统计
                    if row['err'] in ("原交易失败",):
                        self.tx_pay_failed += int(row['count'])
                # 核心支付成功与失败通过 核心的成功失败表去纠正,这里统计的成功不做数
                if str(row['rsp_code']) == "0":
                    if str(row['production_code']) != "SQC":
                        self.tx_succ += int(row['count'])
            if self.tx_sum:
                self.tx_succ_rate = self.tx_succ / self.tx_sum

            # 上面把数据拿完了
            # 下面组织 一条记录
            self.res[row['production_code']] = [
                (datetime.now() + timedelta(days=CONF.OFFSET_DAY)).strftime("%Y-%m-%d"),
                self.zx_sum,
                self.zx_request_failed,
                self.zx_already_regist,
                self.zx_unregist,
                self.zx_ocr_failed,
                self.zx_code_06,
                self.zx_code_04,
                self.zx_succ,
                self.zx_succ_rate or '-',
                self.sx_sum,
                self.sx_request_failed,
                self.sx_succ,
                self.sx_ilog_reject,
                self.sx_net_failed,
                self.sx_antifraud_reject,
                self.sx_succ_rate or '-',
                self.tx_sum,
                self.tx_request_failed,
                self.tx_succ,
                self.tx_antifraud_reject,
                self.tx_ilog_reject,
                self.tx_amtques,
                self.tx_pay_failed,
                self.tx_succ_rate or '-'
            ]
            # 判断是否要换一个产品了
            if self.ch_flag:
                self.inits()
                self.ch_flag = False
        return

    def get_result(self):
        return self.res


class MyExcel:
    def __init__(self, data: dict):
        self.data_struc = []
        for k in data:
            if os.path.isfile(os.path.join(CONF.EXCEL_OUT, f'{k}.xlsx')):
                self.data_struc.append(
                    {"wb": load_workbook(os.path.join(CONF.EXCEL_OUT, f'{k}.xlsx')),
                     "datas": data[k],
                     "wb_name": f'{k}.xlsx'})
        return

    def append_row(self):
        for dic in self.data_struc:
            single_sheet = dic.get('wb').get_sheet_by_name('Sheet1')
            data_row = dic.get('datas')
            for grid in range(1, 1000):
                blank_row = single_sheet.cell(row=grid, column=1)
                if str(single_sheet.cell(row=grid, column=1).value) == str(data_row[0]):
                    for col in range(0, len(data_row)):
                        single_sheet.cell(row=grid, column=col + 1).value = data_row[col]
                    logging.info(f" {dic['wb_name']} 覆盖写入")
                    break
                if not blank_row.value:
                    for col in range(0, len(data_row)):
                        single_sheet.cell(row=blank_row.row, column=col + 1).value = data_row[col]
                    logging.info(f" {dic.get('wb_name')} 首次写入")
                    break
            dic.get('wb').save(os.path.join(CONF.EXCEL_OUT, dic.get("wb_name")))
        return


class FixExcel:
    """ 校对放款成功数据用的 """

    def __init__(self, source_code: str):
        self.channel = CONF.PROD_CHANNEL.get(source_code.upper(), False)
        self.sourcecode = source_code
        if not self.channel:
            raise IndexError("source_code 不存在")
        filename = Utils.fix_filename(datetime.now())
        self.fpath = os.path.join(CONF.EXCEL_TAR_PATH, filename)
        if not os.path.isfile(self.fpath):
            raise FileNotFoundError(f"文件不存在{self.fpath}")
        self.dtframe = pd.read_excel(self.fpath, header=1)
        self.wb = load_workbook(os.path.join(CONF.EXCEL_OUT, self.sourcecode + ".xlsx"))
        self.loan_succ = 0
        self.pay_fail = 0
        self.sheet = self.wb.get_sheet_by_name("Sheet1")
        self.date = ""

    def do_fix(self):
        for col, row in self.dtframe.iterrows():
            if self.channel == row['助代方']:
                if "1" == str(row['status']):
                    self.loan_succ = int(row['笔数'])
                    self.date = str(row['日期'])
                elif "2" == str(row['status']):
                    self.pay_fail = int(row['笔数'])
        for x in range(3, 1000):
            if str(self.sheet.cell(row=x, column=1).value) == self.date:
                self.sheet.cell(row=x, column=20).value = self.loan_succ
                self.sheet.cell(row=x, column=24).value = self.pay_fail
                self.sheet.cell(row=x, column=25).value = self.loan_succ / int(self.sheet.cell(row=x, column=18).value)
                # print(self.sheet.cell(row=x, column=1).value)
                break
        self.wb.save(os.path.join(CONF.EXCEL_OUT, self.sourcecode + ".xlsx"))
        logging.info(self.sourcecode + " 校对完成")


class MyError(Exception):
    def __init_subclass__(cls, **kwargs):
        ...


class DownEmail:
    # 自动下载邮件并保存 到系统的下载文件夹
    def __init__(self, acct=CONF.DOWN_EMAIL_ACCT, pwd=CONF.DOWN_EMAIL_PWD, host=CONF.DOWN_EMAIL_HOST):
        self._acct = acct
        self._pwd = pwd
        self._host = host
        try:
            self._server = poplib.POP3_SSL(self._host, port=poplib.POP3_SSL_PORT)
        except:
            raise MyError("host链接错误")

    def run(self):
        self._server.user(self._acct)
        self._server.pass_(self._pwd)
        msg_count, msg_size = self._server.stat()
        resp, mails, octets = self._server.list()
        for i in range(len(mails), len(mails) - CONF.REC_NUM, -1):  # 遍历最新的N个邮件就够了，不用看太多
            if int(mails[i - 1].decode().split(" ")[-1]) > 595001:
                # 有较大附件的直接跳过
                continue
            resp, lines, octets = self._server.retr(i)
            # 原始文本
            try:
                msg_content = b'\r\n'.join(lines).decode('utf-8')
            except:
                continue
            # 解析邮件
            p = Parser()
            msg = p.parsestr(msg_content)
            try:
                if re.match(".*<(.*)>", msg.get("From")).group(1) != "drbigdata@xwfintech.com":
                    continue
            except:
                continue
                # assert MyError("没有发件人的邮件")
            # 当前日期
            today = datetime.today() + timedelta(days=CONF.OFFSET_DAY + 1)
            today = str(today).replace('-', '').split(" ")[0]
            # 邮件的日期
            try:
                mail_time = time.strptime(msg.get("Received")[-31:-6], '%a, %d %b %Y %H:%M:%S')
            except:
                mail_time = time.strptime(msg.get("Date")[0:24], '%a, %d %b %Y %H:%M:%S')
            mail_time = time.strftime("%Y%m%d", mail_time)
            if mail_time != today:
                # 不是当天的邮件就跳过，这样可以设置下载前几天的附件  判断邮件发送日期是否 = today(依赖conf里面的dayoffset)
                continue
            DownEmail.get_attr(msg)
        return

    @staticmethod
    def decode_str(str_in):
        value, charset = decode_header(str_in)[0]
        if charset:
            value = value.decode(charset)
        return value

    @staticmethod
    def get_attr(msg):
        attr_file = []
        for part in msg.walk():
            # 获取附件名称
            filename = part.get_filename()
            if filename:
                h = Header(filename)
                dh = email.header.decode_header(h)
                filename = dh[0][0]
                if dh[0][1]:
                    filename = DownEmail.decode_str(str(filename, dh[0][1]))
                if os.path.isfile(os.path.join(CONF.DOWN_EMAIL_DIR, filename)):
                    logging.info(f"{filename} 已经下载过了。")
                    return [filename]
                if False:
                    # 这里可以改条件，通过不下载某些文件
                    return
                file_data = part.get_payload(decode=True)
                file_obj = open(os.path.join(CONF.DOWN_EMAIL_DIR, filename), "wb")
                attr_file.append(filename)
                file_obj.write(file_data)
                file_obj.close()
                logging.info(f"{filename} 下载完成。")
        return attr_file


def zipfiles() -> tuple:
    """
    压缩文件
    :return: (内部压缩文件名，外部压缩文件名)
    """
    if not os.path.isdir(CONF.ZIP_OUT):
        os.makedirs(CONF.ZIP_OUT)
    # 内部邮件附件
    f_selfy = zipfile.ZipFile(
        os.path.join(CONF.ZIP_OUT,
                     f'{(datetime.now() + timedelta(days=CONF.OFFSET_DAY)).strftime("%Y-%m-%d")}_助贷流量excel内部统计结果.zip')
        , 'w', zipfile.ZIP_DEFLATED)
    # 内部全量文件
    selfy_files = []
    for f in CONF.SELFY_FILES:
        prod_file = os.path.join(CONF.EXCEL_OUT, f"{f}.xlsx")
        if os.path.isfile(prod_file):
            selfy_files.append(prod_file)

    selfy_files.append(os.path.join(CONF.DOWN_EMAIL_DIR,
                                    f'导流_新农业务数据统计_{(datetime.now() + timedelta(days=CONF.OFFSET_DAY + 1)).strftime("%Y-%m-%d")}.xlsx'))
    selfy_files.append(os.path.join(CONF.DOWN_EMAIL_DIR,
                                    f'助贷_狮桥,360业务数据统计_{(datetime.now() + timedelta(days=CONF.OFFSET_DAY + 1)).strftime("%Y-%m-%d")}.xlsx'))

    # 压缩
    for i in selfy_files:
        file = i.split('/')[-1].split('\\')[-1]
        f_selfy.write(i, file)
    f_selfy.close()

    # 外部邮件附件
    f_out = zipfile.ZipFile(
        os.path.join(CONF.ZIP_OUT,
                     f'{(datetime.now() + timedelta(days=CONF.OFFSET_DAY)).strftime("%Y-%m-%d")}_助贷流量excel统计结果.zip')
        , 'w', zipfile.ZIP_DEFLATED)

    # 对外发送的文件
    out_files = []
    for f in CONF.OUT_FILES:
        prod_file = os.path.join(CONF.EXCEL_OUT, f"{f}.xlsx")
        if os.path.isfile(prod_file):
            out_files.append(prod_file)
    #  如果给对方发清洗后的数据则不用append这两个东西，可以在CONF里面加SQC 和 QH ，新农的没办法，导流与助贷的日志不同
    out_files.append(os.path.join(CONF.DOWN_EMAIL_DIR,
                                  f'导流_新农业务数据统计_{(datetime.now() + timedelta(days=CONF.OFFSET_DAY + 1)).strftime("%Y-%m-%d")}.xlsx'))
    out_files.append(os.path.join(CONF.DOWN_EMAIL_DIR,
                                  f'助贷_狮桥,360业务数据统计_{(datetime.now() + timedelta(days=CONF.OFFSET_DAY + 1)).strftime("%Y-%m-%d")}.xlsx'))

    for i in out_files:
        file = i.split('/')[-1].split('\\')[-1]
        f_out.write(i, file)
    f_out.close()
    logging.info(" 压缩包生成-成功 ")
    return f_selfy.filename.split('\\')[-1], f_out.filename.split('\\')[-1]


def send_email(selfy: str, out: str):
    # 内部邮件
    message_selfy = MIMEMultipart()
    message_selfy.attach(MIMEText(
        "内部的助贷方统计结果\r\n数据源为大数据每日发送的cros_log清洗统计结果.xlsx\r\n与对外的区别为：\r\n对外只发了部分:(HB.XLSX,TPJF.xlsx,LX.xlsx,WX.xlsx助贷XXX.xlsx, 导流xxx.xslx)"))
    attr_selfy = MIMEText(open(os.path.join(CONF.ZIP_OUT, selfy), 'rb').read(), 'base64', 'utf-8')
    attr_selfy['Content-Type'] = 'application/octet-stream'
    attr_selfy.add_header('Content-Disposition', 'attachment', filename=('gbk', '', selfy))
    message_selfy.attach(attr_selfy)
    message_selfy['From'] = formataddr(["Public-数据推送", "publica@xwfintech.com"])
    message_selfy['To'] = Header(';'.join(CONF.EMAIL_SELFY))
    message_selfy['Subject'] = Header(selfy.split('.')[0], 'utf-8')
    try:
        smt_obj = smtplib.SMTP_SSL()
        smt_obj.connect(CONF.MAIL_HOST, 465)
        smt_obj.login(CONF.MAIL_USER, CONF.MAIL_PWD)
        smt_obj.sendmail(CONF.MAIL_USER, CONF.EMAIL_SELFY, message_selfy.as_string())
        logging.info(" 内部邮件发送成功 ")
    except smtplib.SMTPException as e:
        logging.error(f" 内部邮件发送失败-{e} ")

    # 对外邮件
    message_out = MIMEMultipart()
    message_out.attach(MIMEText("上农的各位老师好：\r\n    这是昨日的助贷流量的统计，请查收！\r\n    谢谢！"))
    attr_out = MIMEText(open(os.path.join(CONF.ZIP_OUT, out), 'rb').read(), 'base64', 'utf-8')
    attr_out['Content-Type'] = 'application/octet-stream'
    attr_out.add_header('Content-Disposition', 'attachment', filename=('gbk', '', out))
    message_out.attach(attr_out)
    message_out['From'] = formataddr(["Public-数据推送", "publica@xwfintech.com"])
    message_out['To'] = Header(';'.join(CONF.EMAIL_SRCB))
    message_out['Subject'] = Header(out.split('.')[0], 'utf-8')
    try:
        smt_obj = smtplib.SMTP_SSL()
        smt_obj.connect(CONF.MAIL_HOST, 465)
        smt_obj.login(CONF.MAIL_USER, CONF.MAIL_PWD)
        smt_obj.sendmail(CONF.MAIL_USER, CONF.EMAIL_SRCB, message_out.as_string())
        logging.info(" 对外邮件发送成功 ")
    except smtplib.SMTPException as e:
        logging.error(f" 对外邮件发送失败-{e} ")
    return


# 整体的步骤
def auto_run(fix=("HB",)):
    # 下载当天的所有附件
    down = DownEmail()
    down.run()

    # 处理数据的一套步骤
    filename = Utils.cob_filename(datetime.now())
    move_cros_log(CONF.EXCEL_SOURCE_PATH, CONF.EXCEL_TAR_PATH, filename)
    xl_data = pd.read_excel(os.path.join(CONF.EXCEL_TAR_PATH, filename), header=1)
    xl_data.replace(CONF.PROD_CODE, CONF.PROD_NAME, inplace=True)
    tool = Tool()
    tool.gen_result(xl_data)
    res = tool.get_result()
    for k in res:
        print(k, ":", res[k])
    myexcel = MyExcel(res)
    myexcel.append_row()

    # 校验放款成功数据的一套流程
    fix_org_file = Utils.fix_filename(datetime.now())
    move_cros_log(CONF.EXCEL_SOURCE_PATH, CONF.EXCEL_TAR_PATH, fix_org_file)
    fix_inst = FixExcel(fix[0])
    fix_inst.do_fix()

    # 压缩并邮件发送
    a, b = zipfiles()
    send_email(a, b)


# 下载邮件附件
def sub0_download_email():
    dn = DownEmail()
    dn.run()


# 只生excel的步骤
def sub1_genexcel():
    # 处理数据的一套步骤
    filename = Utils.cob_filename(datetime.now())
    move_cros_log(CONF.EXCEL_SOURCE_PATH, CONF.EXCEL_TAR_PATH, filename)
    xl_data = pd.read_excel(os.path.join(CONF.EXCEL_TAR_PATH, filename), header=1)
    xl_data.replace(CONF.PROD_CODE, CONF.PROD_NAME, inplace=True)
    # print(xl_data)
    tool = Tool()
    tool.gen_result(xl_data)
    res = tool.get_result()
    for k in res:
        print(k, ":", res[k])
    myexcel = MyExcel(res)
    myexcel.append_row()


# 校正放款成功的笔数,fix为哪个产品需要校正
def sub2_fixexcel(fix="HB"):
    # 校验放款成功数据的一套流程
    fix_org_file = Utils.fix_filename(datetime.now())
    move_cros_log(CONF.EXCEL_SOURCE_PATH, CONF.EXCEL_TAR_PATH, fix_org_file)
    fix_inst = FixExcel(fix)
    fix_inst.do_fix()


# 压缩及发送邮件的函数
def sub3_zipandemail():
    a, b = zipfiles()
    send_email(a, b)


if __name__ == "__main__":
    auto_run()
    # sub0_download_email()
    # sub1_genexcel()
    # sub2_fixexcel()
    # sub3_zipandemail()

    ...
