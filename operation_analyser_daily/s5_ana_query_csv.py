import csv
import logging
from datetime import datetime, timedelta
import CONF
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import Utils

logging.basicConfig(format='%(asctime)s|%(levelname)s|%(message)s', level=logging.DEBUG, filename='log_ana.txt')


class WorkBookError(Exception):
    def __init__(self, msg):
        ...


# excel类
class MyWorkBook:
    def __init__(self, filename: str):
        self.workbook_name = filename
        self.workbook = load_workbook(filename)
        self.sheets_names = self.workbook.get_sheet_names
        self.date = (filename.split("-")[1] + "-" + filename.split("-")[2]).split('.')[0]  # 取出里面的日期

    def get_sheet_by_index(self, index: int) -> Worksheet:
        """根据索引获得工作表"""
        if index >= len(self.sheets_names()) or index < 0:
            logging.error("无法根据索引获得工作表.")
            raise WorkBookError("out of index")
        return self.workbook.get_sheet_by_name(self.sheets_names()[index])

    def save(self, filename=None):
        """保存文件"""
        if not filename:
            self.workbook.save(self.workbook_name)
            return
        self.workbook.save(filename)
        return

    def append_row(self, source_row: list, date="2020-01-01"):
        work_sheet = self.get_sheet_by_index(0)
        for x in range(3, 33):
            add_row = work_sheet.cell(row=x, column=1)
            # 判断空行,遇到空行就往里面添加日期及统计的数据
            if not add_row.value:
                source_row.insert(0, f"{date}")
                print(source_row)
                for c in range(0, len(source_row)):
                    work_sheet.cell(row=add_row.row, column=c + 1).value = source_row[c]
                return
            else:
                continue


class NewCsvHandler:
    """一个新的处理CSV的分析器"""

    # 2020-04-12
    def __init__(self, csv_path: str):
        # 放与excel中差不多对应的结果的
        self.result = []
        self.sum_realname = 0  # 征信总笔数
        self.realname_request_failed = 0  # 请求失败,比如姓名大于20啦,地址大于60或者等于0啦
        self.user_already_regist = 0  # 已在上农注册用户数
        self.user_unregist = 0  # 未注册笔数
        self.ocr_fail = 0  # ocr失败数
        self.ocr_succ = 0  # ocr成功笔数
        self.white_box_succ = 0  # 白盒子通过笔数    namelist=0 (太平)
        self.realname_succ = 0  # 其他的就是返回00 就算
        self.white_box_rate = '-'  # 白盒子核准率

        self.sum_credit = 0  # 授信总数
        self.credit_request_failed = 0  # 授信请求出错
        self.credit_succ = 0  # 授信成功
        self.credit_reject = 0  # 信用拒绝
        self.credit_net_fail = 0  # 50003网络错误
        self.credit_antifraud_reject = 0  # 50003 反欺诈拒绝
        self.credit_rate = '-'  # 授信核准率

        self.sum_withdraw = 0  # 提现总数
        self.withdraw_failed = 0  # 50005请求失败,网络错误或者啥的,好像用不上
        self.withdraw_succ = 0  # 提现成功笔数
        self.withdraw_antifraud_reject = 0  # 50005 报反欺诈
        self.withdraw_ilog_reject = 0  # 提现风控拒绝
        self.withdraw_pay_failed = 0  # 支付通道拒绝
        # self.total_amt = 0  # 当日提现总额
        self.withdraw_rate = '-'  # 当日提现成功率

        # 360 专用的字段
        self.withdraw_pay_origin_trans_failed = 0  # 原交易失败
        self.withdraw_pay_notexist_trans = 0  # 无此交易

        # 临时字段
        self.realname_schedu_succ = 0  # 50001 返回值为00的，代表流程调度的成功数量
        self.code_06 = 0  # 50002 反06 白盒查询失败的情况 针对还呗
        self.code_04 = 0  # 50002 反04 征信查询失败的情况

        """用一个csv路径初始化一个解析器
            并从路径中解析出sourceCode
            解析出 rows [[],[],],一行为一个内部列表"""
        self.rows = list()
        try:
            with open(csv_path, "r", encoding="GBK", newline="") as f:
                csv_content = csv.reader(f)
                for t in csv_content:
                    self.rows.append(t)
        except Exception as e:
            try:
                with open(csv_path, "r", encoding="UTF-8", newline="") as f:
                    csv_content = csv.reader(f)
                    for t in csv_content:
                        self.rows.append(t)
            except Exception as e:
                logging.error(e)
        finally:
            sourcecode = csv_path.split("-")[0].split("/")[-1]
        self.sourceCode = sourcecode

    def gen_result(self):
        """生成一个列表[,,,],对应相应产品的excel的行里面的数字,
        其实搞个sqlite会很方便,但算了吧"""
        # 针对 360-QH 进行的额外补充
        if self.sourceCode == "QH":
            self.gen_360_result()
            return
            # pass
        # 太平金服有白盒核准率什么的,要单独算
        for row in self.rows:  # [[],[],...] 取出内部的 []
            # 征信
            if row[0] == "realnameauth":
                self.sum_realname += int(row[-1])
                if row[1] == "00":
                    self.realname_schedu_succ += int(row[-1])
                if row[1] == "01":
                    self.realname_request_failed += int(row[-1])
                if row[1] == "03":
                    self.user_already_regist += int(row[-1])
                #     未注册用户数 = 总量-请求失败的-已注册的
                self.user_unregist = self.sum_realname - self.realname_request_failed - self.user_already_regist
            elif row[0] == "realnameauth_query":
                if row[1] == "08":
                    self.ocr_fail += int(row[-1])
                if row[1] == "00":
                    self.realname_succ += int(row[-1])
                if row[-2] == "0":
                    #  太平白盒子通过的
                    self.white_box_succ = int(row[-1])
                if row[1] == "06":
                    self.code_06 += int(row[-1])
                if row[1] == "04":
                    self.code_04 += int(row[-1])
                #  白盒子通过率 = 白盒子name list 0/ocr成功的
                if self.ocr_succ != 0 and self.sourceCode.upper() == "TPJF":
                    self.white_box_rate = self.white_box_succ / self.ocr_succ
                elif self.sum_realname != 0 and self.sourceCode.upper() != "TPJF":
                    self.white_box_rate = self.realname_succ / self.user_unregist
                elif self.sum_realname == 0:
                    self.white_box_rate = "-"  # 缺省值
                # ocr成功的 = realnameauth为00的 - ocr失败
                self.ocr_succ = self.realname_schedu_succ - self.ocr_fail
            # 授信
            elif row[0] == "credit":
                self.sum_credit += int(row[-1])
                if self.sourceCode.upper() == "HB":
                    # 还呗有了结果回调后就不去查 结果了,所以只能以这个为准
                    if row[2] == "unknow" and len(row[1]) > 2:
                        self.credit_net_fail += int(row[-1])
                    if row[-2] == "antifraud reject":
                        self.credit_antifraud_reject += int(row[-1])
                    if row[1] == "01" and row[-2] != "antifraud reject":
                        self.credit_request_failed += int(row[-1])
                    if row[1] == "00" and row[2] == "reject":
                        self.credit_reject += int(row[-1])
                    if row[1] == "00" and row[2] == "success":
                        self.credit_succ += int(row[-1])
                    if self.sum_credit != 0:
                        self.credit_rate = self.credit_succ / self.sum_credit
                else:
                    if row[2] == "unknow" and len(row[1]) > 2:
                        self.credit_net_fail += int(row[-1])
                    if row[-2] == "antifraud reject":
                        self.credit_antifraud_reject += int(row[-1])
                    if row[1] == "01" and row[-2] != "antifraud reject":
                        self.credit_request_failed += int(row[-1])
            elif row[0] == "credit_query":
                if self.sourceCode.upper() != "HB":
                    # 还呗不能用这个 log_type 方式进行统计了
                    if row[1] == "02":
                        self.credit_reject += int(row[-1])
                    if row[1] == "01":
                        self.credit_succ += int(row[-1])
                    if self.sum_credit != 0:
                        self.credit_rate = self.credit_succ / self.sum_credit
                    else:
                        self.credit_rate = "-"  # 总数为0的话，缺省值为 -
            # 提现
            elif row[0] == "withdraw":
                self.sum_withdraw += int(row[-1])
                if row[1] == "01" and row[-2] == "网络错误":
                    self.withdraw_failed += int(row[-1])
                if row[1] == "01" and row[-2] == "antifraud reject":
                    self.withdraw_antifraud_reject += int(row[-1])
                if len(str(row[1])) > 2:
                    self.withdraw_failed += int(row[-1])
                if row[3] == "request param invalid" or "timeout" in row[3]:
                    self.withdraw_failed += int(row[-1])
                # 还呗可能要单独处理一下
                if self.sourceCode.upper() == "HB":
                    ...
                else:
                    ...
            elif row[0] == "withdraw_query":
                if row[-2] == "放款失败——ilog拒绝":
                    self.withdraw_ilog_reject += int(row[-1])
                if row[-2] == "call loan_mng service timeout":
                    self.withdraw_pay_failed += int(row[-1])  # 这个归类的支付错误吧
                if self.sourceCode.upper() != "HB":
                    # 还呗不能用这个 log_type 方式进行统计了
                    if row[2] == "处理成功":
                        self.withdraw_succ += int(row[-1])
                    if len(row[-2]) > 15:
                        self.withdraw_pay_failed += int(row[-1])
                    if row[-2] == "原交易失败":
                        self.withdraw_pay_failed += int(row[-1])
                    if self.sum_withdraw != 0:
                        self.withdraw_rate = self.withdraw_succ / self.sum_withdraw
                    else:
                        self.withdraw_rate = "-"  # 缺省值 这样可以不计入平均值计算

        #  将结果组织到
        self.result = []
        self.result.append(self.sum_realname or 0)
        self.result.append(self.realname_request_failed or 0)
        self.result.append(self.user_already_regist or 0)
        self.result.append(self.user_unregist or 0)
        self.result.append(self.ocr_fail or 0)

        if self.sourceCode in ("HB", "LX", "QH"):
            # 还呗把ocr成功替换为征信查询白盒子失败
            self.result.append(self.code_06)
            self.result.append(self.code_04)
        else:
            # TPJF
            self.result.append(self.ocr_succ or 0)

        # 后面的 insert 位 使这个的排列与excel的表头一致
        if self.sourceCode.upper() == "TPJF":
            self.result.insert(7, self.white_box_succ)
        else:
            self.result.insert(7, self.realname_succ)

        self.result.append(self.white_box_rate)
        self.result.append(self.sum_credit or 0)
        self.result.append(self.credit_request_failed or 0)
        self.result.append(self.credit_succ or 0)
        self.result.append(self.credit_reject or 0)
        self.result.append(self.credit_net_fail or 0)
        self.result.append(self.credit_antifraud_reject or 0)
        self.result.append(self.credit_rate)

        self.result.append(self.sum_withdraw or 0)
        self.result.append(self.withdraw_failed or 0)
        self.result.append(self.withdraw_succ or 0)
        self.result.append(self.withdraw_antifraud_reject or 0)
        self.result.append(self.withdraw_ilog_reject or 0)
        self.result.append(self.withdraw_pay_failed or 0)
        # self.result.append("总额") # 由于这一列不太重要,每次都要手动去看,故省去,
        self.result.append(self.withdraw_rate)

    def gen_360_result(self):
        """针对 360 专门生成结果列表, 360的统计方式与excel表头与另外三个产品大不相同故单独列出来"""
        for row in self.rows:
            if row[0] == "withdraw":
                if row[3] == "禁止注册":
                    self.user_already_regist += int(row[-1])
                if row[3] == "授信限制":
                    self.credit_reject += int(row[-1])
                if row[3] == "反欺诈拒绝":
                    self.withdraw_antifraud_reject += int(row[-1])
            if row[0] == "withdraw_query":
                if row[1] == "success":
                    self.withdraw_succ += int(row[-1])
                if row[2] == "授信拒绝":
                    self.credit_reject += int(row[-1])
                if row[2] == "其它错误":
                    self.withdraw_ilog_reject += int(row[-1])
                if row[2] == "原交易失败":
                    self.withdraw_pay_origin_trans_failed += int(row[-1])
                if row[2] == "无此交易":
                    self.withdraw_pay_notexist_trans += int(row[-1])

        self.result = []
        self.result.append(self.withdraw_succ)
        self.result.append(self.user_already_regist)
        self.result.append(self.credit_reject)
        self.result.append(self.withdraw_antifraud_reject)
        self.result.append(self.withdraw_ilog_reject)
        self.result.append(self.withdraw_pay_origin_trans_failed)
        self.result.append(self.withdraw_pay_notexist_trans)

    def get_result(self):
        return self.result


def main(csv_path: str, workbook_path: str, source_code=None):
    ins = NewCsvHandler(csv_path)
    if not source_code:
        ins.gen_result()
    elif source_code == "QH":
        ins.gen_360_result()
    res = ins.get_result()

    wb = MyWorkBook(workbook_path)
    date_tpm = csv_path.split("/")[-1].split("-")[1]  # 20200101
    date = date_tpm[0:4] + "-" + date_tpm[4:6] + "-" + date_tpm[6:]
    wb.append_row(res, date)
    wb.save()


if __name__ == '__main__':
    def xm(date: str = None):
        source_codes = ("HB", "TPJF", "LX",)
        # source_codes = ("TPJF",)
        # 每日分析结果用的文件名
        if not date:
            date = (datetime.now() + timedelta(days=CONF.DATE_OFFSET)).strftime("%Y%m%d")
        files = Utils.gen_res_file_path(date, source_codes)
        # excel文件名用的日期
        simple_date = (datetime.now() + timedelta(days=CONF.DATE_OFFSET)).strftime("%Y-%m")

        for f in files:
            if "TPJF" in f:
                main(f, f"excels/TPJF-{simple_date}.xlsx")
            if "HB" in f:
                main(f, f"excels/HB-{simple_date}.xlsx")
            if "LX" in f:
                main(f, f"excels/LX-{simple_date}.xlsx")
            if "QH" in f:
                main(f, f"excels/QH-{simple_date}.xlsx", source_code="")


    xm()
    # for x in range(20200501,20200519):
    #     xm(str(x))
